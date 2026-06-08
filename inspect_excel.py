import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
path = sys.argv[1]
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Indices municipales"]

# Row 5 (first data row) - cols 0-12
print("Row 5 first 13 cols:")
row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
for i in range(13):
    print(f"  col {i}: {repr(row5[i])}")

print(f"\nRow 5 col 10-39 (ODS indices):")
for i in range(10, 40):
    print(f"  col {i}: {repr(row5[i])}")

print(f"\nMax row: {ws.max_row}")

# Count non-None rows in col 0
count = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    if row[0] is not None:
        count += 1
print(f"Non-empty data rows: {count}")
